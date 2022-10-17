from readline import read_history_file
import pandas as pd
import numpy as np
import io
import matplotlib.pyplot as plt
from matplotlib.gridspec import GridSpec
from matplotlib.figure import Figure
import seaborn as sns
from datetime import date
import os

import openpyxl
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Font, Border, Side, PatternFill, Alignment
from openpyxl.formatting.rule import CellIsRule
from openpyxl.drawing.image import Image
from openpyxl.utils.dataframe import dataframe_to_rows


class ExcelReport():
    def __init__(self, tableFormat:dict=None):
        self.cur_sheet = None
        self.cur_content = None
        self.ws = None
        self.tableFormat = tableFormat

    def create(self, sheets_data:dict, path:str):
        if not isinstance(sheets_data, dict):
            raise ValueError('"data" parameter should be dictionary')

        folder = os.path.split(path)[0]
        if not os.path.exists(folder):
            os.makedirs(folder)

        wb = openpyxl.Workbook()
        for sheet, contents in sheets_data.items():
            if not isinstance(sheet, str):
                raise ValueError('"data" parameter should contains string as key')
            else:
                self.cur_sheet = sheet
            if not hasattr(contents, '__iter__'):
                raise ValueError('"data" parameter should contains iteriable type as value')

            self.ws = wb.create_sheet(sheet)
            for content in contents:
                self.cur_content = content
                data = self.cur_content.get('data')
                asTable = self.cur_content.get('asTable')
                formatTable = self.cur_content.get('formatTable')
                importIndex = self.cur_content.get('importIndex')

                if isinstance(data, pd.DataFrame):
                    if data.shape[0] == 0:
                        continue
                    self.write_pandas(data, index=importIndex)
                    if asTable:
                        self.convertto_Table(data)
                    if formatTable:
                        self.format_ReportTable(data)
                
                elif isinstance(data, Figure):
                    self.drawFigure(data)

        del wb['Sheet']
        wb.save(path)

        wb = openpyxl.load_workbook(path)
        wb.save(path)
    
    def check_parameters(self, parameters, type_):
        values = []
        for content_key in parameters:
            content_value = self.cur_content.get(content_key)
            if isinstance(content_value, type_):
                values.append(content_value)
            else:
                raise ValueError(f'{content_key} parameter should be {type_}.')
        return values
    
    def write_pandas(self, data, index=False, header=True):
        self.startrow, self.startcol = self.check_parameters(('startrow', 'startcol'), int)
        rows = dataframe_to_rows(data, index=index, header=header)
        for rowidx, row in enumerate(rows):
            for colidx, value in enumerate(row):
                wrow, wcol = (rowidx + self.startrow + 1), (colidx + self.startcol + 1)
                self.ws.cell(row=wrow, column=wcol).value = value
    
    def convertto_Table(self, df):
        nrows, ncols = df.shape
        cell_range = f'A1:{self.get_ColumnLetter(ncols)}{nrows+1}'
        tab = Table(displayName="prediction", ref=cell_range)
        style = TableStyleInfo(name="TableStyleLight1", showFirstColumn=False,
                            showLastColumn=False, showRowStripes=True, showColumnStripes=False)
        tab.tableStyleInfo = style
        self.ws.add_table(tab)

        # change columns width
        for num in range(1, ncols+1):
            ind = self.get_ColumnLetter(num)
            self.ws.column_dimensions[ind].width = 18
        return
    
    def get_ColumnLetter(self, col_num):
        letter1, letter2 = col_num // 26, col_num % 26
        
        if letter2 == 0:
            letter1, letter2 = -1, 26
        
        if letter1 > 0:
            ColumnLetter = f"{chr(ord('@') + letter1)}{chr(ord('@') + letter2)}"
        else:
            ColumnLetter = chr(ord('@') + letter2)
        return ColumnLetter
    
    def format_ReportTable(self, df):
        self.index_start = (df.index.shape[0], len(df.index.names))
        self.column_start = (len(df.columns.names) + 1, df.columns.shape[0])
        self.tabvalue_start = df.shape

        if self.tableFormat is None:    
            self.generate_tableFormat()

        self.TitleFormat()
        self.tabvalueFormat()
        self.indexFormat()
        self.columnFormat()       

    def TitleFormat(self):
        title_name, *_ = self.check_parameters(['title_name'], str)
        
        if len(title_name) > 0:
            row_title, col_title = self.startrow-2, self.startcol + 1
            title = self.ws.cell(row=row_title, column=col_title)
            title.value = title_name
            title.font = self.tableFormat['TitleFont']
            title.alignment = self.tableFormat['TitleAlignment']
            self.ws.merge_cells(
                start_row=row_title, start_column=col_title,
                end_row=row_title, end_column=col_title+self.index_start[1]+self.column_start[1]
            )
    
    def generate_tableFormat(self):
        self.tableFormat = {
            'MedGreyFill': PatternFill('solid', start_color='A6A6A6'),
            'LightGreyFill' : PatternFill('solid', start_color='BFBFBF'),
            'LightGreenFill' : PatternFill('solid', start_color='CCEBCE', end_color='CCEBCE'),
            'DarkGreenFont' : Font(color='286017'),
            'ValueBD' : Side(border_style='thin'),
            'HeaderFont' : Font(bold=True),
            'IndexFont' : Font(bold=True),
            'TitleAlignment' : Alignment(horizontal='center'),
            'TitleFont' : Font(bold=True, size=18)
        }
        ValueBD = self.tableFormat['ValueBD']
        self.tableFormat['ValueBorder'] = Border(top=ValueBD, right=ValueBD, bottom=ValueBD, left=ValueBD)

    def tabvalueFormat(self):
        for row_idx in range(1, self.tabvalue_start[0] + 1):
            fill_ = PatternFill()
            for col_idx in range(1, self.tabvalue_start[1] + 1):
                r = self.startrow + self.column_start[0] + row_idx
                c = self.startcol + self.index_start[1] + col_idx
                cell_ = self.ws.cell(row=r, column=c)

                if col_idx == 1:
                    prev_cell_ = cell_.offset(column=-1).value
                    #TODO add prec_cell_ value as parameter
                    if prev_cell_ == 'Итого сеть':
                        fill_ = self.tableFormat['MedGreyFill']
                #TODO add number format as parameter
                cell_.number_format = '0.00%'
                cell_.border = self.tableFormat['ValueBorder']
                cell_.fill = fill_

            self.conditionFormat(col_idx, row_idx, r, c)
    
    def conditionFormat(self, col_idx, row_idx, r, c):
        # conditional formating values
        try:
            cell_range = (f'{self.get_ColumnLetter(c-col_idx+1)}{r-row_idx+1}:{self.get_ColumnLetter(c)}{r}')
            rule = CellIsRule(
                operator='greaterThanOrEqual', formula=['0.8'],
                stopIfTrue=False, fill=self.tableFormat['LightGreenFill'],
                font=self.tableFormat['DarkGreenFont']
            )
            self.ws.conditional_formatting.add(cell_range, rule)
        except TypeError:
            raise ValueError(f' Can bot to apply conditional formating to call range.\n'
                'Check parameters: c = {c}, r = {r}, row_idx = {row_idx}, col_idx = {col_idx}')
    
    def indexFormat(self):
        # format report index
        for col_idx in range(1, self.index_start[1] + 1):
            prev_cell = None
            for row_idx in range(0, self.index_start[0] + 1):
                r = self.startrow + self.column_start[0] + row_idx
                c = self.startcol + col_idx

                cell_ = self.ws.cell(row=r, column=c)
                cell_.font = self.tableFormat['IndexFont']
                cell_.border = self.tableFormat['ValueBorder']
                # use another color for last index column
                if col_idx == self.index_start[1] and cell_.value != 'Итого сеть':
                    cell_.fill = self.tableFormat['LightGreyFill']
                else:
                    cell_.fill = self.tableFormat['MedGreyFill']
                
                cell_val = cell_.value
                if prev_cell is None:
                    prev_cell = cell_val
                    merge_start = r

                if ((cell_val is not None) and (cell_val != prev_cell)):
                    if prev_cell is not None:
                        self.ws.merge_cells(
                            start_row=merge_start, start_column=c,
                            end_row=r-1, end_column=c
                        )
                    prev_cell = cell_val
                    merge_start = r
                
                if (row_idx == self.index_start[0]):
                    self.ws.merge_cells(
                            start_row=merge_start, start_column=c,
                            end_row=r, end_column=c
                        )

        ind = chr(ord('@') + (self.startcol + col_idx))        
        self.ws.column_dimensions[ind].width = 30
    
    def columnFormat(self):
        # format report header
        for row_idx in range(1, self.column_start[0] + 1):
            prev_cell = None
            for col_idx in range(1, self.column_start[1] + 1):
                r = self.startrow + row_idx
                c = self.startcol + self.index_start[1] + col_idx
                cell_ = self.ws.cell(row=r, column=c)
                cell_.font = self.tableFormat['HeaderFont']
                cell_.fill = self.tableFormat['MedGreyFill']
                cell_.border = self.tableFormat['ValueBorder']

                cell_val = cell_.value
                if prev_cell is None:
                    prev_cell = cell_val
                    merge_start = c

                if ((cell_val is not None) and (cell_val != prev_cell)):
                    if prev_cell is not None:
                        self.ws.merge_cells(
                            start_row=r, start_column=merge_start,
                            end_row=r, end_column=c-1
                        )
                    prev_cell = cell_val
                    merge_start = c
                
                if (col_idx == self.column_start[1]):
                    self.ws.merge_cells(
                            start_row=r, start_column=merge_start,
                            end_row=r, end_column=c
                        )
    
    def drawFigure(self, fig):
        self.startrow, self.startcol = self.check_parameters(('startrow', 'startcol'), int)
        imgdata=io.BytesIO()
        fig.savefig(imgdata, format='png')
        ColumnLetter = self.get_ColumnLetter(self.startcol)
        exl_cell = f'{ColumnLetter}{self.startrow}'
        self.ws.add_image(Image(imgdata), exl_cell)
